import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill

def compare_gdhi(combined_authority_file, output_area_file):
    # Load the Excel files
    ca_df = pd.read_excel(combined_authority_file)
    oa_df = pd.read_excel(output_area_file)

    # Filter the rows where metrics are 'GDHI'
    ca_gdhi = ca_df[ca_df['metrics'] == 'GDHI']
    oa_gdhi = oa_df[oa_df['metrics'] == 'GDHI']

    # Prepare a dictionary to store the results
    results = {
        'Combine authority': [],
        'Year': [],
        'Combined Authority GDHI': [],
        'Sum of Output Areas GDHI': [],
        'Difference': [],
        'Percentage Difference': []
    }

    # Iterate over the years 2005, 2006, 2007
    for year in ['2005', '2006', '2007']:
        for ca in ca_gdhi['Combine authority'].unique():
            ca_value = ca_gdhi.loc[ca_gdhi['Combine authority'] == ca, year].values[0]
            oa_values = oa_gdhi.loc[oa_gdhi['Combine authority'] == ca, year].sum()
            
            difference = ca_value - oa_values
            percent_difference = (difference / ca_value) * 100 if ca_value != 0 else 0

            # Append the results
            results['Combine authority'].append(ca)
            results['Year'].append(year)
            results['Combined Authority GDHI'].append(ca_value)
            results['Sum of Output Areas GDHI'].append(oa_values)
            results['Difference'].append(difference)
            results['Percentage Difference'].append(percent_difference)

    # Convert results dictionary to DataFrame
    results_df = pd.DataFrame(results)

    # Write the results to a new Excel file
    output_file = 'comparison_results.xlsx'
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        results_df.to_excel(writer, sheet_name='Results', index=False)
        
        # Apply conditional formatting
        workbook = writer.book
        worksheet = writer.sheets['Results']

        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        for row in range(2, len(results_df) + 2):
            percent_diff = results_df['Percentage Difference'][row - 2]
            cell = worksheet[f'F{row}']
            if abs(percent_diff) < 1:
                cell.fill = green_fill
            else:
                cell.fill = red_fill

    print(f'Results have been saved to {output_file}')

# Example usage:
combined_authority_file = 'combined_authority.xlsx'
output_area_file = 'output_area.xlsx'

compare_gdhi(combined_authority_file, output_area_file)
