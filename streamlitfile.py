import pandas as pd
import streamlit as st
from openpyxl.styles import PatternFill

# Function to compare GDHI values
def compare_gdhi(ca_df, oa_df):
    # Filter the rows where metrics are 'GDHI'
    ca_gdhi = ca_df[ca_df['metrics'] == 'GDHI']
    oa_gdhi = oa_df[oa_df['metrics'] == 'GDHI']

    # Prepare a dictionary to store the results
    results = {
        'Combine authority': [],
        'Year': [],
        'Combined Authority GDHI': [],
        'Sum of Output Areas GDHI': [],
        'Difference': [],
        'Percentage Difference': []
    }

    # Iterate over the years 2005, 2006, 2007
    for year in ['2005', '2006', '2007']:
        for ca in ca_gdhi['Combine authority'].unique():
            ca_value = ca_gdhi.loc[ca_gdhi['Combine authority'] == ca, year].values[0]
            oa_values = oa_gdhi.loc[oa_gdhi['Combine authority'] == ca, year].sum()
            
            difference = ca_value - oa_values
            percent_difference = (difference / ca_value) * 100 if ca_value != 0 else 0

            # Append the results
            results['Combine authority'].append(ca)
            results['Year'].append(year)
            results['Combined Authority GDHI'].append(ca_value)
            results['Sum of Output Areas GDHI'].append(oa_values)
            results['Difference'].append(difference)
            results['Percentage Difference'].append(percent_difference)

    # Convert results dictionary to DataFrame
    results_df = pd.DataFrame(results)

    return results_df

# Streamlit app
st.title("GDHI Quality Assurance Check")

# Sidebar for file upload and table name input
st.sidebar.header("Upload Files")
ca_file = st.sidebar.file_uploader("Upload Combined Authority Excel File", type=["xlsx"])
oa_file = st.sidebar.file_uploader("Upload Output Area Excel File", type=["xlsx"])

# Call to action button
if st.sidebar.button("Run QA Check"):
    if ca_file and oa_file:
        ca_df = pd.read_excel(ca_file, header=1)
        oa_df = pd.read_excel(oa_file, header=1)
        
        # Ensure columns are correctly interpreted as strings
        ca_df.columns = ca_df.columns.map(str)
        oa_df.columns = oa_df.columns.map(str)

        # Perform the QA check
        results_df = compare_gdhi(ca_df, oa_df)

        # Save results to an Excel file
        output_file = 'comparison_results.xlsx'
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            results_df.to_excel(writer, sheet_name='Results', index=False)
            
            # Apply conditional formatting
            workbook = writer.book
            worksheet = writer.sheets['Results']

            green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

            for row in range(2, len(results_df) + 2):
                percent_diff = results_df['Percentage Difference'][row - 2]
                cell = worksheet[f'F{row}']
                if abs(percent_diff) < 1:
                    cell.fill = green_fill
                else:
                    cell.fill = red_fill

        # Display success message
        st.success("QA Check completed and results saved to 'comparison_results.xlsx'.")

        # Display the results DataFrame
        st.write("Comparison Results:")
        st.dataframe(results_df)

        # Provide a download link for the Excel file
        with open(output_file, "rb") as file:
            btn = st.download_button(
                label="Download Results",
                data=file,
                file_name=output_file,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    else:
        st.error("Please upload both Excel files.")
